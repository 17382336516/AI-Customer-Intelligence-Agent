from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
import numpy as np
import pandas as pd
from sklearn.cluster import KMeans
from sklearn.metrics import silhouette_score
from sklearn.preprocessing import StandardScaler

from ..config import settings

FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "user_id": ("user_id", "userid", "user", "uid", "用户id", "用户_id", "客户id"),
    "amount": ("amount", "price", "gmv", "sales", "金额", "消费金额", "订单金额", "实付"),
    # 收入字段：仅作为「输入属性 / 用户画像字段」，系统不负责按收入筛选人群。
    "income": (
        "income", "annual_income", "yearly_income", "year_income", "salary", "年收入",
        "年收入和", "收入", "月收入", "月薪", "收入层级", "income_level", "income_band",
    ),
    "category": ("category", "cate", "品类", "分类", "商品分类", "一级品类"),
    "product": ("product", "product_name", "item", "sku_name", "商品", "商品名", "商品名称"),
    "event_time": (
        "event_time",
        "order_time",
        "created_at",
        "timestamp",
        "date",
        "时间",
        "消费时间",
        "下单时间",
        "日期",
    ),
    "order_id": ("order_id", "transaction_id", "订单id", "交易id", "流水号"),
    "status": ("status", "order_status", "状态", "订单状态"),
}

# ---------------------------------------------------------------------------
# 品类兜底匹配规则（FALLBACK ONLY）
# ---------------------------------------------------------------------------
# 设计原则：
#   1. 这套规则「只用于兜底」——只有当原始 category 缺失、为空或完全无法解析
#      （纯自由文本、不含任何结构化层级）时，才用关键词去猜。
#   2. 结构化 category（如 electronics.smartphone / computers.notebook /
#      appliances.kitchen.refrigerators）会被 normalize_category 直接保留，
#      不经过下面的关键词猜测，避免二次猜测导致归类失真。
#   3. 关键词必须「精确」，禁止过宽词：
#      - 禁止单独使用 "智能" / "手机" 这类词，因为「智能手表 / 智能家居 / 智能设备」
#        不一定属于手机；"手机" 只与明确手机语义共存时命中。
#      - smartphone 仅保留 iphone / smartphone / phone / 手机 等明确命中词，
#        不再包含 "智能"（单独）与厂商泛化词（samsung / xiaomi 等已移除，
#        因为它们也会命中耳机、电视等其他品类）。
#   4. 兜底匹配「只基于 product 字段」（自由文本），绝不再把原始 category 混进来。
# ---------------------------------------------------------------------------
CATEGORY_PATTERNS: dict[str, tuple[str, ...]] = {
    "smartphone": ("iphone", "smartphone", "mobilephone", "mobile phone", "手机"),
    "electronics": ("earphone", "headphone", "speaker", "camera", "tablet", "电子", "数码"),
    "computer": ("notebook", "laptop", "desktop", "电脑", "笔记本"),
    "appliance": ("refrigerator", "washer", "vacuum", "oven", "家电", "冰箱", "洗衣机"),
    "apparel": ("apparel", "shoes", "clothes", "服饰", "鞋服", "鞋子"),
    "furniture": ("furniture", "bedroom", "sofa", "cabinet", "家具", "沙发"),
    "kids": ("toys", "carriage", "儿童", "母婴", "玩具"),
    "auto": ("tires", "car accessory", "汽车", "车品", "车载"),
    "sport": ("fitness", "运动", "健身"),
    "coffee": ("espresso", "cafe", "咖啡", "咖啡杯"),
    "tea": ("teapot", "奶茶", "茶饮", "茶具"),
    "pet": ("puppy", "kitten", "萌宠", "宠物", "猫", "狗"),
    "travel": ("luggage", "passport", "旅行", "旅游", "机票", "酒店", "出行"),
    "kitchen": ("baking", "recipe", "餐厨", "烘焙", "厨房"),
    "home": ("candle", "frame", "家居", "居家", "日用"),
    "christmas": ("christmas", "xmas", "圣诞", "节日"),
    "vintage": ("vintage", "retro", "复古"),
    "gift": ("ornament", "ribbon", "礼品", "礼盒"),
}

# 结构化 category 的顶层前缀 -> 中文展示名。
# 用于把 electronics.smartphone / computers.notebook 这类层级值转成友好中文，
# 不改变底层标准化 key。
STRUCTURED_CATEGORY_CN: dict[str, str] = {
    "electronics": "数码电子",
    "computers": "电脑办公",
    "appliances": "家电",
    "apparel": "服饰鞋履",
    "furniture": "家居家具",
    "kids": "母婴儿童",
    "auto": "汽车用品",
    "sport": "运动健身",
    "travel": "旅行",
    "tea": "茶饮",
    "coffee": "咖啡",
    "pet": "萌宠",
    "kitchen": "餐厨",
    "home": "家居日用",
    "christmas": "节日布置",
    "vintage": "复古审美",
    "gift": "礼品",
}


def category_display(category: Any) -> str:
    """把标准化品类 key 转成展示用中文名。

    - 兜底 key（smartphone/electronics/...）走 CATEGORY_CN。
    - 结构化层级 key（electronics.smartphone）取其顶层前缀的中文名 + 细分后缀，
      例如 electronics.smartphone -> "数码电子·智能手机"。
    """
    key = str(category).strip()
    if key in CATEGORY_CN:
        return CATEGORY_CN[key]
    if "." in key:
        top, *rest = key.split(".")
        top_cn = STRUCTURED_CATEGORY_CN.get(top, top)
        sub = "·".join(rest)
        return f"{top_cn}·{sub}" if sub else top_cn
    if key in STRUCTURED_CATEGORY_CN:
        return STRUCTURED_CATEGORY_CN[key]
    return "综合兴趣"
# 例如：electronics.smartphone / computers.notebook / appliances.kitchen.refrigerators
_STRUCTURED_CATEGORY_RE = re.compile(r"[A-Za-z0-9]+(?:\.[A-Za-z0-9]+)+")


def _looks_structured(value: str) -> bool:
    """判断原始 category 是否为结构化分类（可直接采用，无需关键词猜测）。"""
    text = str(value).strip()
    if not text or text.lower() in {"nan", "none", "null", ""}:
        return False
    # 点分层级：electronics.smartphone
    if _STRUCTURED_CATEGORY_RE.search(text):
        return True
    # 含层级分隔或明确的英文品类标识（带下划线/连字符的英文 token）：computers_notebook
    if re.search(r"[A-Za-z0-9]+(?:[_-][A-Za-z0-9]+)+", text):
        return True
    return False


def _category_from_text(value: str) -> str:
    """兜底：仅对自由文本做关键词匹配。结构化 category 不应走这里。"""
    text = str(value).strip().lower()
    if not text or text in {"nan", "none", "null", ""}:
        return "other"
    for category, patterns in CATEGORY_PATTERNS.items():
        if any(pattern in text for pattern in patterns):
            return category
    return "other"


def normalize_category(category: Any, product: Any = "") -> str:
    """品类标准化：结构化原始 category 优先，关键词匹配仅作兜底。

    优先级：
      1. 若原始 category 是结构化分类（点分层级 / 英文层级名），直接保留原值。
      2. 否则，若 category 为空 / 无法解析，改用 product 文本做关键词兜底。
      3. 若 product 也无法解析，返回 "other"。

    这样可避免对已结构化的 category 二次猜测（例如把
    electronics.smartphone 之外的字段误并入 smartphone）。
    """
    raw = "" if category is None else str(category).strip()
    if _looks_structured(raw):
        return raw
    # 原始 category 不是结构化值：尝试用 category 本身做关键词兜底。
    if raw and raw.lower() not in {"nan", "none", "null", ""}:
        guessed = _category_from_text(raw)
        if guessed != "other":
            return guessed
    # 仍无法解析：退回用 product 关键词兜底（绝不把原始 category 与 product 拼接后猜测）。
    product_text = "" if product is None else str(product)
    return _category_from_text(product_text)

CATEGORY_CN: dict[str, str] = {
    "smartphone": "智能手机",
    "electronics": "数码电子",
    "computer": "电脑办公",
    "appliance": "家电",
    "apparel": "服饰鞋履",
    "furniture": "家居家具",
    "kids": "母婴儿童",
    "auto": "汽车用品",
    "sport": "运动健身",
    "travel": "旅行",
    "tea": "茶饮",
    "coffee": "咖啡",
    "pet": "萌宠",
    "kitchen": "烘焙餐厨",
    "home": "家居日用",
    "christmas": "节日布置",
    "vintage": "复古审美",
    "gift": "礼品",
    "high_value": "高价值",
    "general_interest": "综合兴趣",
}


def build_cluster_statistics(
    cluster_features: pd.DataFrame,
    *,
    overall_monetary: float,
    overall_frequency: float,
    overall_recency: float,
) -> dict[str, Any]:
    """汇总单个聚类（cluster）的真实统计特征，供名称、画像与策略复用。"""
    monetary = float(cluster_features["monetary"].mean())
    frequency = float(cluster_features["frequency"].mean())
    recency = float(cluster_features["recency"].mean())
    category_diversity = float(cluster_features["category_diversity"].mean())
    weekend_share = float(cluster_features["weekend_share"].mean())
    afternoon_share = float(cluster_features["afternoon_share"].mean())
    purchase_count = float(cluster_features["purchase_count"].mean())
    cart_count = float(cluster_features["cart_count"].mean())
    view_count = float(cluster_features["view_count"].mean())
    event_count = float(cluster_features["event_count"].mean())
    active_days = float(cluster_features["active_days"].mean())
    category_concentration = float(cluster_features["category_concentration"].mean()) if "category_concentration" in cluster_features else 0.0
    unique_category_count = float(cluster_features["unique_category_count"].mean()) if "unique_category_count" in cluster_features else 0.0
    weighted_purchase_value = float(cluster_features["weighted_purchase_value"].mean()) if "weighted_purchase_value" in cluster_features else 0.0

    # 主品类基于「金额 × 行为权重」的加权得分，而非简单浏览次数。
    # 这里直接基于标准化后真实品类列的 *_weighted / *_share（动态 key），
    # 不再限定于固定兜底集合，从而忠实反映原始品类分布。
    weighted_columns = [col for col in cluster_features.columns if col.endswith("_weighted")]
    weighted_totals = {
        col[: -len("_weighted")]: float(cluster_features[col].sum())
        for col in weighted_columns
    }
    weighted_total_all = sum(weighted_totals.values()) or 1.0
    category_distribution: dict[str, float] = {}
    for col in weighted_columns:
        category = col[: -len("_weighted")]
        share_col = f"{category}_share"
        if share_col in cluster_features.columns:
            mean_share = float(cluster_features[share_col].mean())
            if mean_share > 0:
                category_distribution[category] = round(mean_share, 4)
    main_category = (
        max(weighted_totals, key=weighted_totals.get)
        if any(v > 0 for v in weighted_totals.values())
        else None
    )
    main_category_ratio = (
        round(weighted_totals.get(main_category, 0.0) / weighted_total_all, 4)
        if main_category
        else 0.0
    )
    dominant_category = main_category
    dominant_ratio = main_category_ratio
    overall_monetary = overall_monetary or 1.0
    return {
        "average_spend": round(monetary, 2),
        "overall_average_spend": round(overall_monetary, 2),
        "spend_ratio": round(monetary / overall_monetary, 2),
        "average_frequency": round(frequency, 2),
        "overall_average_frequency": round(overall_frequency, 2),
        "average_recency": round(recency, 1),
        "overall_average_recency": round(overall_recency, 1),
        "average_category_count": round(category_diversity, 2),
        "weekend_ratio": round(weekend_share, 4),
        "afternoon_ratio": round(afternoon_share, 4),
        "purchase_count": round(purchase_count, 2),
        "cart_count": round(cart_count, 2),
        "view_count": round(view_count, 2),
        "event_count": round(event_count, 2),
        "average_active_days": round(active_days, 2),
        "category_distribution": category_distribution,
        "main_category": main_category,
        "main_category_ratio": round(main_category_ratio, 4),
        "dominant_category": dominant_category,
        "dominant_category_ratio": round(dominant_ratio, 4),
        "category_concentration": round(category_concentration, 4),
        "unique_category_count": round(unique_category_count, 2),
        "weighted_purchase_value": round(weighted_purchase_value, 2),
        # 聚类解释别名：直接对应"为什么这些用户属于一类"所需字段。
        "cluster_size": int(len(cluster_features)),
        "avg_spend": round(monetary, 2),
        "avg_frequency": round(frequency, 2),
        "top_category": main_category,
        "category_diversity": round(category_diversity, 2),
        "user_count": int(len(cluster_features)),
    }


def build_cluster_profile(stats: dict[str, Any]) -> tuple[str, list[str]]:
    """依据聚类统计特征生成人群名称与画像，不套用任何预设模板。"""
    overall_monetary = float(stats.get("overall_average_spend", 0.0)) or 1.0
    overall_frequency = float(stats.get("overall_average_frequency", 0.0)) or 1.0
    overall_recency = float(stats.get("overall_average_recency", 0.0)) or 1.0
    monetary = float(stats.get("average_spend", 0.0))
    frequency = float(stats.get("average_frequency", 0.0))
    recency = float(stats.get("average_recency", 0.0))
    category_diversity = float(stats.get("average_category_count", 0.0))
    weekend = float(stats.get("weekend_ratio", 0.0))
    afternoon = float(stats.get("afternoon_ratio", 0.0))
    dominant = stats.get("main_category") or stats.get("dominant_category")
    dominant_ratio = float(stats.get("main_category_ratio", 0.0) or stats.get("dominant_category_ratio", 0.0))

    modifiers: list[str] = []
    if monetary >= overall_monetary * 1.1:
        modifiers.append("高客单")
    if frequency >= overall_frequency * 1.1:
        modifiers.append("高频")
    if recency <= overall_recency * 0.8:
        modifiers.append("近期活跃")
    if category_diversity >= 3:
        modifiers.append("跨品类")
    modifier_text = "·".join(modifiers)

    if dominant and dominant_ratio >= 0.15:
        cn = category_display(str(dominant))
        name = f"{modifier_text}{cn}人群" if modifier_text else f"{cn}偏好人群"
    else:
        name = f"综合兴趣{'·' + modifier_text if modifier_text else ''}人群"

    spend_cmp = "高于" if monetary > overall_monetary else "接近" if monetary >= overall_monetary * 0.9 else "低于"
    freq_cmp = "高于" if frequency > overall_frequency else "接近" if frequency >= overall_frequency * 0.9 else "低于"
    spend_ratio_local = monetary / overall_monetary
    ratio_disp = f"{spend_ratio_local:.1f}" if spend_ratio_local >= 0.1 else f"{spend_ratio_local:.2f}"
    features: list[str] = [
        f"平均消费 ¥{monetary:.0f}，{spend_cmp}整体平均 ¥{overall_monetary:.0f}（{ratio_disp} 倍）",
        f"平均购买 {frequency:.1f} 次，{freq_cmp}整体 {overall_frequency:.1f} 次",
        f"最近一次消费平均在 {recency:.0f} 天前（整体 {overall_recency:.0f} 天）",
        f"平均覆盖 {category_diversity:.1f} 个品类",
    ]
    if dominant:
        cn = category_display(str(dominant))
        features.append(f"主品类为{cn}，群内加权贡献占 {dominant_ratio:.0%}（金额×行为权重）")
    if weekend >= 0.3:
        features.append(f"周末消费占 {weekend:.0%}")
    elif weekend <= 0.15:
        features.append(f"工作日消费为主（周末占 {weekend:.0%}）")
    if afternoon >= 0.4:
        features.append(f"下午时段（13–18 点）消费占 {afternoon:.0%}")
    return name, features[:6]


def build_recommended_strategy(stats: dict[str, Any]) -> str:
    """基于聚类统计特征生成数据驱动的推荐策略方向。"""
    overall_monetary = float(stats.get("overall_average_spend", 0.0)) or 1.0
    overall_frequency = float(stats.get("overall_average_frequency", 0.0)) or 1.0
    overall_recency = float(stats.get("overall_average_recency", 0.0)) or 1.0
    monetary = float(stats.get("average_spend", 0.0))
    frequency = float(stats.get("average_frequency", 0.0))
    recency = float(stats.get("average_recency", 0.0))
    dominant = stats.get("dominant_category")
    spend_ratio = monetary / overall_monetary

    parts: list[str] = []
    if spend_ratio >= 1.5:
        parts.append("针对高客单特征配置专属满额权益")
    elif spend_ratio >= 1.1:
        parts.append("在常规客单基础上叠加品类券提升转化")
    if frequency >= overall_frequency * 1.1:
        parts.append("用复购返券或攒钱任务锁定高频节奏")
    if recency <= overall_recency * 0.8:
        parts.append("趁近期活跃窗口推送首发或限时权益")
    if dominant:
        cn = category_display(str(dominant))
        parts.append(f"围绕{cn}做定向品类券与内容承接")
    if not parts:
        parts.append("可用主题集合继续识别兴趣方向，再分配轻量权益测试")
    return "；".join(parts) + "。"


FIELD_ALIASES["user_id"] += ("customerid", "customer_id")
FIELD_ALIASES["product"] += ("description", "product_description")
FIELD_ALIASES["event_time"] += ("invoicedate", "invoice_date")
FIELD_ALIASES["order_id"] += ("invoiceno", "invoice_no", "invoice")
FIELD_ALIASES["quantity"] = ("quantity", "qty", "count")
FIELD_ALIASES["unit_price"] = ("unitprice", "unit_price", "price_each")


def _normal(value: str) -> str:
    return re.sub(r"[\s_\-]+", "", str(value).strip().lower())


def detect_mapping(columns: list[str]) -> dict[str, str]:
    normalized = {_normal(column): column for column in columns}
    mapping: dict[str, str] = {}
    for canonical, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            if _normal(alias) in normalized:
                mapping[canonical] = normalized[_normal(alias)]
                break
    return mapping


def read_dataset(path: str | Path) -> pd.DataFrame:
    path = Path(path)
    if path.suffix.lower() == ".csv":
        for encoding in ("utf-8-sig", "utf-8", "gb18030"):
            try:
                return pd.read_csv(path, encoding=encoding)
            except UnicodeDecodeError:
                continue
        raise ValueError("无法识别 CSV 编码，请转换为 UTF-8 后重试。")
    if path.suffix.lower() in {".xlsx", ".xls"}:
        return pd.read_excel(path)
    raise ValueError("仅支持 CSV、XLSX 和 XLS 文件。")


def _category_from_text(value: str) -> str:
    text = str(value).strip().lower()
    for category, patterns in CATEGORY_PATTERNS.items():
        if any(pattern in text for pattern in patterns):
            return category
    return "other"


def estimate_row_count(path: str | Path) -> int | None:
    path = Path(path)
    if path.suffix.lower() == ".csv":
        for encoding in ("utf-8-sig", "utf-8", "gb18030"):
            try:
                with path.open("r", encoding=encoding, errors="strict") as handle:
                    return max(0, sum(1 for _ in handle) - 1)
            except UnicodeDecodeError:
                continue
        return None
    if path.suffix.lower() in {".xlsx", ".xls"}:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            return max(0, (workbook.active.max_row or 1) - 1)
        finally:
            workbook.close()
    return None


def read_dataset_preview(path: str | Path, rows: int = 20000) -> tuple[pd.DataFrame, int]:
    path = Path(path)
    total_rows = estimate_row_count(path)
    if path.suffix.lower() == ".csv":
        for encoding in ("utf-8-sig", "utf-8", "gb18030"):
            try:
                frame = pd.read_csv(path, encoding=encoding, nrows=rows)
                return frame, total_rows if total_rows is not None else len(frame)
            except UnicodeDecodeError:
                continue
        raise ValueError("Unable to detect CSV encoding. Please convert it to UTF-8.")
    if path.suffix.lower() in {".xlsx", ".xls"}:
        frame = pd.read_excel(path, nrows=rows)
        return frame, total_rows if total_rows is not None else len(frame)
    raise ValueError("Only CSV, XLSX and XLS files are supported.")


def build_quality_report(raw: pd.DataFrame) -> dict[str, Any]:
    mapping = detect_mapping([str(column) for column in raw.columns])
    issues: list[dict[str, Any]] = []
    required = ("user_id", "amount", "event_time")
    has_derived_amount = "quantity" in mapping and "unit_price" in mapping
    missing_required = [
        field
        for field in required
        if field not in mapping and not (field == "amount" and has_derived_amount)
    ]
    for field in missing_required:
        issues.append(
            {
                "code": "missing_required_field",
                "severity": "error",
                "message": f"缺少必填字段：{field}",
                "field": field,
            }
        )
    if "category" not in mapping and "product" not in mapping:
        issues.append(
            {
                "code": "missing_category_source",
                "severity": "error",
                "message": "category 与 product 至少需要一个，用于识别人群偏好。",
                "field": "category",
            }
        )

    missing_rates = {
        canonical: round(float(raw[column].isna().mean()), 4)
        for canonical, column in mapping.items()
    }
    if "amount" not in mapping and has_derived_amount:
        quantity = pd.to_numeric(raw[mapping["quantity"]], errors="coerce")
        unit_price = pd.to_numeric(raw[mapping["unit_price"]], errors="coerce")
        missing_rates["amount"] = round(float((quantity.isna() | unit_price.isna()).mean()), 4)
        mapping["amount"] = "__derived_amount__"
    user_count = int(raw[mapping["user_id"]].nunique()) if "user_id" in mapping else 0
    usable_mask = pd.Series(True, index=raw.index)
    for field in required:
        if field == "amount" and mapping.get("amount") == "__derived_amount__":
            quantity = pd.to_numeric(raw[mapping["quantity"]], errors="coerce")
            unit_price = pd.to_numeric(raw[mapping["unit_price"]], errors="coerce")
            usable_mask &= quantity.notna() & unit_price.notna()
        elif field in mapping:
            usable_mask &= raw[mapping[field]].notna()
    usable_rows = int(usable_mask.sum()) if len(raw) else 0

    category_coverage = 0.0
    if "category" in mapping:
        category_coverage = float(raw[mapping["category"]].notna().mean())
    elif "product" in mapping:
        category_coverage = float(raw[mapping["product"]].notna().mean())

    score = 100
    score -= len(missing_required) * 25
    if "category" not in mapping and "product" not in mapping:
        score -= 20
    score -= int(sum(rate > 0.2 for rate in missing_rates.values()) * 5)
    if len(raw) < settings.min_analyzable_rows:
        score -= 15
        issues.append(
            {
                "code": "small_sample",
                "severity": "warning",
                "message": f"样本少于 {settings.min_analyzable_rows} 行，结论稳定性有限。",
                "field": None,
            }
        )
    if user_count < 5:
        score -= 15
        issues.append(
            {
                "code": "few_users",
                "severity": "warning",
                "message": "独立用户少于 5 人，不适合进行稳定分群。",
                "field": "user_id",
            }
        )
    score = max(0, min(100, score))
    can_analyze = not missing_required and (
        "category" in mapping or "product" in mapping
    ) and usable_rows > 0

    return {
        "row_count": len(raw),
        "usable_row_count": usable_rows,
        "user_count": user_count,
        "analyzability_score": score,
        "field_mapping": mapping,
        "missing_rates": missing_rates,
        "category_coverage": round(category_coverage, 4),
        "issues": issues,
        "can_analyze": can_analyze,
    }


# 行为事件权重：购买 > 加购 > 浏览，用于加权主品类识别（金额 × 行为权重）。
EVENT_WEIGHTS: dict[str, float] = {
    "purchase": 5.0,
    "cart": 3.0,
    "view": 1.0,
}
_DEFAULT_EVENT_WEIGHT = 3.0  # 未提供 event_type 时按中性权重处理（介于浏览与购买之间）。


def _event_weight(event_type: Any) -> float:
    if not event_type:
        return _DEFAULT_EVENT_WEIGHT
    key = str(event_type).strip().lower()
    if key in EVENT_WEIGHTS:
        return EVENT_WEIGHTS[key]
    if "购买" in key or "下单" in key or "paid" in key or "order" in key:
        return EVENT_WEIGHTS["purchase"]
    if "加购" in key or "cart" in key:
        return EVENT_WEIGHTS["cart"]
    if "浏览" in key or "view" in key or "访问" in key:
        return EVENT_WEIGHTS["view"]
    return _DEFAULT_EVENT_WEIGHT


def clean_dataset(raw: pd.DataFrame, mapping: dict[str, str]) -> tuple[pd.DataFrame, dict[str, int]]:
    derived_amount = mapping.get("amount") == "__derived_amount__"
    rename_map = {
        source: target
        for target, source in mapping.items()
        if source != "__derived_amount__"
    }
    renamed = raw.rename(columns=rename_map).copy()
    stats = {"input_rows": len(renamed), "duplicates_removed": 0, "invalid_removed": 0}

    renamed["user_id"] = renamed["user_id"].astype("string").str.strip()
    if derived_amount:
        renamed["amount"] = (
            pd.to_numeric(renamed["quantity"], errors="coerce")
            * pd.to_numeric(renamed["unit_price"], errors="coerce")
        )
    else:
        renamed["amount"] = pd.to_numeric(renamed["amount"], errors="coerce")
    # 收入字段：仅作用户画像属性，系统不据此筛选人群。
    # 解析失败/缺失统一置为 NaN（表示「样本未提供收入」，下游优雅降级）。
    if "income" in renamed.columns:
        renamed["income"] = pd.to_numeric(renamed["income"], errors="coerce")
    else:
        renamed["income"] = np.nan
    renamed["event_time"] = pd.to_datetime(renamed["event_time"], errors="coerce")
    if "event_type" in renamed.columns:
        renamed["event_type"] = renamed["event_type"].astype(str).str.strip()
    else:
        renamed["event_type"] = "purchase"

    before = len(renamed)
    if "order_id" in renamed.columns and "product" in renamed.columns:
        subset = ["order_id", "product"]
    elif "order_id" in renamed.columns:
        subset = ["order_id"]
    else:
        subset = [
            "user_id",
            "amount",
            "event_time",
        ]
    renamed = renamed.drop_duplicates(subset=subset)
    stats["duplicates_removed"] = int(before - len(renamed))

    invalid = renamed["user_id"].isna() | renamed["event_time"].isna() | renamed["amount"].isna()
    invalid |= renamed["amount"] <= 0
    if "status" in renamed.columns:
        invalid |= renamed["status"].astype(str).str.lower().str.contains(
            "refund|cancel|退款|取消", regex=True, na=False
        )
    stats["invalid_removed"] = int(invalid.sum())
    renamed = renamed.loc[~invalid].copy()

    if "product" not in renamed.columns:
        renamed["product"] = ""

    # 品类标准化：结构化原始 category 优先，关键词匹配仅作兜底（基于 product）。
    # 不再把 category 与 product 拼接后做关键词猜测，避免对已结构化数据二次猜测。
    renamed["normalized_category"] = [
        normalize_category(cat, prod)
        for cat, prod in zip(renamed["category"], renamed["product"], strict=False)
    ]
    renamed["weekday"] = renamed["event_time"].dt.dayofweek
    renamed["hour"] = renamed["event_time"].dt.hour
    renamed["event_weight"] = renamed["event_type"].map(_event_weight)
    return renamed, stats


def build_user_features(cleaned: pd.DataFrame) -> pd.DataFrame:
    reference_date = cleaned["event_time"].max().normalize() + pd.Timedelta(days=1)
    grouped = cleaned.groupby("user_id", observed=True)
    features = grouped.agg(
        recency=("event_time", lambda values: (reference_date - values.max()).days),
        frequency=("amount", "size"),
        monetary=("amount", "sum"),
        avg_order_value=("amount", "mean"),
        active_days=("event_time", lambda values: values.dt.date.nunique()),
        category_diversity=("normalized_category", "nunique"),
        weekend_share=("weekday", lambda values: float((values >= 5).mean())),
        afternoon_share=("hour", lambda values: float(((values >= 13) & (values <= 18)).mean())),
        purchase_count=("event_type", lambda values: int((values.str.lower() == "purchase").sum())),
        cart_count=("event_type", lambda values: int((values.str.lower() == "cart").sum())),
        view_count=("event_type", lambda values: int((values.str.lower() == "view").sum())),
        event_count=("event_type", "size"),
    )
    # 加权行为特征：直接基于「标准化后的品类」(normalized_category) 的真实取值，
    # 不再把数据强行映射到固定的兜底 key 集合（CATEGORY_PATTERNS）。
    # 这样主品类统计会忠实反映原始结构化分布（如 electronics.smartphone
    # 不会被错误并入 smartphone），也不会出现单一兜底品类被放大的失真。
    # 列名沿用 {品类}_weighted / {品类}_share，便于下游 concentration 等指标复用。
    weighted_long = cleaned.assign(
        _w=cleaned["amount"] * cleaned["event_weight"]
    )
    pivot_weighted = weighted_long.pivot_table(
        index="user_id",
        columns="normalized_category",
        values="_w",
        aggfunc="sum",
        fill_value=0.0,
    )
    count_long = cleaned.groupby(["user_id", "normalized_category"], observed=True).size().unstack(fill_value=0)
    count_shares = count_long.div(count_long.sum(axis=1).replace(0, np.nan), axis=0).fillna(0.0)

    weighted_matrix = pivot_weighted.reindex(features.index).fillna(0.0)
    share_matrix = count_shares.reindex(features.index).fillna(0.0)
    # 一次性拼接品类列，避免逐列 insert 导致的 DataFrame 碎片化（品类数较多时明显）。
    cat_cols: dict[str, pd.Series] = {}
    for category in weighted_matrix.columns:
        cat_cols[f"{category}_weighted"] = weighted_matrix[category]
        cat_cols[f"{category}_share"] = share_matrix.get(
            category, pd.Series(0.0, index=features.index)
        )
    features = pd.concat([features, pd.DataFrame(cat_cols, index=features.index)], axis=1)

    features = features.fillna(0)

    # ---- 用户价值特征（区分"看得多"与"真正买得多"）----
    # 仅购买事件计入真实购买金额与订单数；浏览/加购不计入实际成交。
    cleaned_purchase = cleaned[cleaned["event_type"].astype(str).str.lower() == "purchase"]
    purchase_agg = cleaned_purchase.groupby("user_id", observed=True).agg(
        total_purchase_amount=("amount", "sum"),
        purchase_order_count=("amount", "size"),
    )
    features["total_purchase_amount"] = features.index.map(
        purchase_agg["total_purchase_amount"]
    ).fillna(0.0)
    features["purchase_order_count"] = features.index.map(
        purchase_agg["purchase_order_count"]
    ).fillna(0).astype(int)
    features["average_purchase_value"] = (
        features["total_purchase_amount"] / features["purchase_order_count"].replace(0, np.nan)
    ).fillna(0.0)

    # ---- 高价值行为：金额 × 行为权重（购买 5 / 加购 3 / 浏览 1）----
    features["weighted_purchase_value"] = (
        cleaned.assign(weighted=cleaned["amount"] * cleaned["event_weight"])
        .groupby("user_id", observed=True)["weighted"]
        .sum()
        .reindex(features.index)
        .fillna(0.0)
    )

    # ---- 消费集中度：消费金额在各类目间的分布均匀度（赫芬达尔指数 0~1）----
    # 1 = 完全集中到单一品类；越接近 0 = 消费越分散。
    weighted_row_totals = weighted_matrix.sum(axis=1).replace(0, np.nan)
    weighted_shares = weighted_matrix.div(weighted_row_totals, axis=0).fillna(0.0)
    features["category_concentration"] = (weighted_shares ** 2).sum(axis=1).round(4)

    # ---- 品类偏好多样性：用户实际发生消费（加权>0）的品类数量 ----
    features["unique_category_count"] = (weighted_matrix > 0).sum(axis=1).astype(int)

    # 别名：聚类特征与下游解释统一使用 recency_days（值等同于 recency）。
    features["recency_days"] = features["recency"]

    # ---- 收入画像（输入属性，非系统推断）----
    # 收入本来就是用户属性，同一用户多行应相同；用均值即可，缺失为 NaN。
    if "income" in cleaned.columns:
        income_by_user = cleaned.groupby("user_id", observed=True)["income"].mean()
        features["income"] = features.index.map(income_by_user)
    else:
        features["income"] = np.nan

    features = features.reset_index()
    return features


def select_best_cluster_count(matrix: pd.DataFrame, *, min_k: int = 2, max_k: int = 8) -> tuple[int, float]:
    """在 K=min_k..max_k 范围内，选择 silhouette_score 最高的聚类数。

    返回 (最优 K, 对应的轮廓系数)。若无法计算（样本过少或方差为 0），优雅降级返回 (max(2, 样本数), 0.0)。
    """
    n_samples = len(matrix)
    upper = min(max_k, max(min_k, n_samples - 1))
    if upper < min_k or n_samples < 2:
        return max(2, min(min_k, n_samples)), 0.0
    if float(np.var(matrix)) == 0.0:
        return max(2, n_samples), 0.0

    best_k = upper
    best_score = -1.0
    for k in range(min_k, upper + 1):
        try:
            labels = KMeans(n_clusters=k, random_state=42, n_init=10).fit_predict(matrix)
            if len(set(labels)) < 2:
                continue
            score = silhouette_score(matrix, labels)
        except ValueError:
            continue
        if score > best_score:
            best_score = float(score)
            best_k = k
    return best_k, round(best_score, 4)


# 送入 KMeans 的特征筛选结果（feature selection）。
# 设计原则：保留商业可解释且区分度高的人群维度，剔除噪声/冗余字段，
# 从而让聚类边界更清晰、结果更易解释，而不是靠调 K 拉高 silhouette。
CLUSTERING_FEATURES: tuple[str, ...] = (
    "total_purchase_amount",   # 用户实际购买金额（价值）
    "purchase_order_count",    # 实际购买订单数（购买频次）
    "average_purchase_value",  # 平均购买金额（客单价）
    "recency_days",            # 最近消费间隔（活跃度）
    "category_concentration",  # 消费集中度（品类聚焦程度）
    "unique_category_count",   # 品类偏好多样性（覆盖品类数）
    "weighted_purchase_value", # 高价值行为：金额 × 行为权重
    "active_days",             # 活跃天数
)


def user_main_category(features: pd.DataFrame) -> pd.Series:
    """计算每个用户的主品类（基于金额×行为权重的加权贡献最高者）。

    返回 user_id -> 主品类 key 的 Series。加权列名为 {品类}_weighted。
    若某用户所有品类加权均为 0，返回 None（视为无明确偏好）。
    """
    weighted_cols = [c for c in features.columns if c.endswith("_weighted")]
    if not weighted_cols:
        return pd.Series(None, index=features.index, dtype=object)
    wm = features[weighted_cols].copy()
    wm.columns = [c[: -len("_weighted")] for c in wm.columns]
    # 每个用户加权最高的品类
    main = wm.idxmax(axis=1)
    main = main.where(wm.max(axis=1) > 0, other=None)
    return main


def _parent_category(category: str) -> str | None:
    """返回结构化品类的父级（一级），例如 electronics.smartphone -> electronics。

    非结构化兜底 key（smartphone/electronics）无父级，返回 None。
    """
    if not category or "." not in category:
        return None
    return category.split(".")[0]


# 最细粒度品类的友好中文名（贴近用户认知的"商品类别"）。
# 未列出的细分路径回退为「一级中文 + 英文 leaf」以保证可解释。
CATEGORY_LEAF_CN: dict[str, str] = {
    "electronics.smartphone": "智能手机",
    "electronics.video.tv": "电视",
    "electronics.clocks": "钟表",
    "electronics.tablet": "平板",
    "electronics.audio.headphone": "耳机",
    "electronics.audio.subwoofer": "低音炮",
    "electronics.audio.music_tools.piano": "电子琴",
    "computers.notebook": "笔记本电脑",
    "computers.desktop": "台式机",
    "computers.peripherals.mouse": "鼠标",
    "computers.peripherals.keyboard": "键盘",
    "computers.peripherals.printer": "打印机",
    "appliances.kitchen.dishwasher": "洗碗机",
    "appliances.kitchen.refrigerators": "冰箱",
    "appliances.kitchen.washer": "洗衣机",
    "appliances.kitchen.microwave": "微波炉",
    "appliances.kitchen.meat_grinder": "绞肉机",
    "appliances.iron": "电熨斗",
    "appliances.environment.vacuum": "吸尘器",
    "appliances.environment.air_heater": "电暖器",
    "furniture.bedroom.bed": "床",
    "furniture.living_room.sofa": "沙发",
    "furniture.living_room.cabinet": "柜子",
    "apparel.shoes": "鞋",
    "apparel.shoes.keds": "帆布鞋",
    "auto.accessories.alarm": "汽车报警器",
    "auto.accessories.videoregister": "车载记录仪",
    "auto.accessories.player": "车载播放器",
    "auto.accessories.compressor": "车载气泵",
    "kids.toys": "玩具",
    "kids.carriage": "婴儿车",
    "kids.skates": "童鞋轮滑",
}


# 英文品类词 -> 中文（用于兜底未显式映射的细粒度品类，保证名称可读）。
EN_WORD_CN: dict[str, str] = {
    "smartphone": "智能手机", "phone": "手机", "telephone": "电话", "tablet": "平板",
    "tv": "电视", "video": "影音", "audio": "音频", "headphone": "耳机",
    "subwoofer": "低音炮", "acoustic": "音响", "microphone": "麦克风",
    "music_tools": "乐器", "piano": "电子琴", "camera": "相机", "photo": "相机",
    "projector": "投影仪", "clocks": "钟表", "computer": "电脑", "notebook": "笔记本",
    "desktop": "台式机", "ebooks": "电子书", "peripherals": "外设", "mouse": "鼠标",
    "keyboard": "键盘", "monitor": "显示器", "printer": "打印机", "components": "配件",
    "cooler": "散热器", "cpu": "处理器", "hdd": "硬盘", "memory": "内存",
    "motherboard": "主板", "power_supply": "电源", "videocards": "显卡",
    "appliances": "家电", "kitchen": "厨房", "oven": "烤箱", "dishwasher": "洗碗机",
    "refrigerators": "冰箱", "washer": "洗衣机", "microwave": "微波炉",
    "meat_grinder": "绞肉机", "blender": "搅拌机", "coffee_machine": "咖啡机",
    "coffee_grinder": "磨豆机", "juicer": "榨汁机", "kettle": "水壶", "grill": "烤架",
    "hob": "灶具", "hood": "油烟机", "mixer": "料理机", "steam_cooker": "蒸锅",
    "toster": "烤面包机", "iron": "电熨斗", "ironing_board": "熨衣板",
    "environment": "环境", "vacuum": "吸尘器", "air_conditioner": "空调",
    "air_heater": "电暖器", "fan": "风扇", "water_heater": "热水器",
    "personal": "个护", "hair_cutter": "理发器", "massager": "按摩器", "scales": "体重秤",
    "sewing_machine": "缝纫机", "furniture": "家具", "bedroom": "卧室", "bed": "床",
    "blanket": "毯子", "pillow": "枕头", "living_room": "客厅", "sofa": "沙发",
    "cabinet": "柜子", "chair": "椅", "table": "桌", "bathroom": "卫浴", "bath": "浴缸",
    "toilet": "马桶", "universal": "通用", "light": "灯具", "apparel": "服饰",
    "shoes": "鞋", "ballet_shoes": "芭蕾鞋", "keds": "帆布鞋", "moccasins": "软鞋",
    "sandals": "凉鞋", "slipons": "懒人鞋", "costume": "戏服", "dress": "连衣裙",
    "jeans": "牛仔裤", "jumper": "针织衫", "scarf": "围巾", "shirt": "衬衫",
    "trousers": "长裤", "tshirt": "T恤", "underwear": "内衣", "bag": "箱包",
    "accessories": "配件", "alarm": "报警器", "compressor": "气泵", "player": "播放器",
    "radar": "雷达", "videoregister": "记录仪", "winch": "绞盘", "auto": "汽车",
    "kids": "母婴", "carriage": "婴儿车", "dolls": "玩偶", "fmcg": "母婴用品",
    "diapers": "纸尿裤", "skates": "轮滑", "swing": "秋千", "toys": "玩具",
    "sport": "运动", "bicycle": "自行车", "ski": "滑雪", "snowboard": "单板",
    "tennis": "网球", "trainer": "运动鞋", "construction": "建材", "tools": "工具",
    "generator": "发电机", "country_yard": "庭院", "hammok": "吊床",
    "stationery": "文具", "cartrige": "墨盒",
}


def _translate_path(key: str) -> str:
    """把点分品类路径逐段翻译成中文，例如 appliances.kitchen.oven -> 家电·厨房·烤箱。"""
    parts = key.split(".")
    return "·".join(EN_WORD_CN.get(p, p) for p in parts)


def category_leaf_display(category: Any) -> str:
    """最细粒度品类的中文展示名，例如 electronics.smartphone -> '智能手机'。"""
    key = str(category).strip()
    if key in CATEGORY_LEAF_CN:
        return CATEGORY_LEAF_CN[key]
    if "." in key:
        top, *rest = key.split(".")
        top_cn = STRUCTURED_CATEGORY_CN.get(top, top)
        # 已显式映射的二级路径（如 electronics.smartphone）优先；否则逐段翻译。
        sub_cn = _translate_path(".".join(rest)) if rest else ""
        return f"{top_cn}·{sub_cn}" if sub_cn else top_cn
    return category_display(key)


def build_category_segment_name(main_category: str | None, main_ratio: float, *, electronics_only: bool, stats: dict[str, Any] | None = None) -> str:
    """按主品类 + 真实统计特征生成人群名称。

    规则（业务要求）：
      - 主品类占比 Top1 >= 35%：使用最细粒度品类名（如 智能手机 / 洗碗机）。
      - 主品类占比 < 35%：禁止单一商品类命名，使用聚合称谓：
          若全部来自 electronics 一级 -> "综合电子消费用户"
          否则 -> "多品类消费用户" / "跨品类消费用户"。
      - 名称结合真实统计特征（收入画像、客单价、频次、活跃度），不写死标签：
          例如「高收入智能手机购买者」「品质生活咖啡消费人群」「高频旅行消费人群」。
    """
    stats = stats or {}
    overall_monetary = float(stats.get("overall_average_spend", 0.0)) or 1.0
    overall_frequency = float(stats.get("overall_average_frequency", 0.0)) or 1.0
    overall_recency = float(stats.get("overall_average_recency", 0.0)) or 1.0
    monetary = float(stats.get("average_spend", 0.0))
    frequency = float(stats.get("average_frequency", 0.0))
    recency = float(stats.get("average_recency", 0.0))
    avg_income = stats.get("average_income")
    avg_income = float(avg_income) if isinstance(avg_income, (int, float)) and not pd.isna(avg_income) else None

    if not main_category or main_ratio < 0.35:
        if electronics_only:
            base = "综合电子消费用户"
        else:
            base = "多品类消费用户"
        if avg_income is not None:
            return f"高收入{base}" if avg_income >= 300000 else base
        return base

    leaf = category_leaf_display(main_category)
    # 修饰词：全部来自真实统计特征。
    modifiers: list[str] = []
    if avg_income is not None:
        # 收入门槛：>= 30 万视为高收入（与常见「年收入30万以上」语义一致）。
        if avg_income >= 300000:
            modifiers.append("高收入")
        elif avg_income < 120000:
            modifiers.append("大众收入")
    if monetary >= overall_monetary * 1.1:
        modifiers.append("高客单")
    if frequency >= overall_frequency * 1.1:
        modifiers.append("高频")
    if recency <= overall_recency * 0.8:
        modifiers.append("近期活跃")
    # 「品质生活」用于客单与频次都偏高的非数码综合品类，体现消费品质。
    if (
        not electronics_only
        and monetary >= overall_monetary * 1.05
        and frequency >= overall_frequency * 0.95
        and not modifiers
    ):
        modifiers.append("品质生活")
    modifier_text = "".join(modifiers)
    return f"{modifier_text}{leaf}消费人群"


# ---------------------------------------------------------------------------
# 收入画像（输入属性，非系统推断）
# ---------------------------------------------------------------------------
def build_income_profile(features: pd.DataFrame) -> dict[str, Any]:
    """从用户级 income 字段统计收入画像。

    收入是输入数据自带的用户属性，系统不筛选、不推断。
    若数据未提供 income，返回 available=False，下游优雅降级（不虚构）。
    """
    if "income" not in features.columns or features["income"].isna().all():
        return {
            "available": False,
            "average_income": None,
            "median_income": None,
            "high_income_share": None,
            "income_bands": [],
            "note": "样本未提供收入字段，收入画像不可用。",
        }
    series = features["income"].dropna()
    if series.empty:
        return {
            "available": False,
            "average_income": None,
            "median_income": None,
            "high_income_share": None,
            "income_bands": [],
            "note": "样本未提供收入字段，收入画像不可用。",
        }
    avg = float(series.mean())
    median = float(series.median())
    high_share = float((series >= 300000).mean())
    # 收入分层（用于画像展示）：大众 <12万 / 中产 12-30万 / 高净值 >=30万。
    bands = [
        {"label": "大众收入 (<12万)", "share": round(float((series < 120000).mean()), 4)},
        {"label": "中产收入 (12-30万)", "share": round(float(((series >= 120000) & (series < 300000)).mean()), 4)},
        {"label": "高净值 (≥30万)", "share": round(float((series >= 300000).mean()), 4)},
    ]
    return {
        "available": True,
        "average_income": round(avg, 2),
        "median_income": round(median, 2),
        "high_income_share": round(high_share, 4),
        "income_bands": bands,
        "note": "收入来自输入数据（用户属性），仅供画像参考，不用于筛选人群。",
    }


# ---------------------------------------------------------------------------
# 消费趋势分析（第一部分模块）：基于整体输入数据的真实统计
# ---------------------------------------------------------------------------
def build_consumption_trend(
    cleaned: pd.DataFrame,
    features: pd.DataFrame,
    segments: list[dict[str, Any]],
) -> dict[str, Any]:
    """分析输入用户整体消费趋势（真实统计，禁止模板生成）。

    输出 overall_consumption_insight：
      - top_categories: 按加权贡献 Top N 品类
      - category_spend_distribution: 各品类金额贡献占比
      - category_user_distribution: 各品类覆盖用户占比
      - recent_activity_pattern: 最近消费时段 / 活跃节奏
    """
    if cleaned.empty:
        return {
            "available": False,
            "top_categories": [],
            "category_spend_distribution": [],
            "category_user_distribution": [],
            "recent_activity_pattern": {},
        }

    # 1) 金额贡献（spend）按品类聚合
    spend_by_cat = cleaned.groupby("normalized_category", observed=True)["amount"].sum()
    total_spend = float(spend_by_cat.sum()) or 1.0
    spend_dist = (
        (spend_by_cat / total_spend)
        .sort_values(ascending=False)
        .head(6)
    )
    category_spend_distribution = [
        {
            "category": str(cat),
            "category_cn": category_display(cat),
            "spend": round(float(spend_by_cat[cat]), 2),
            "spend_share": round(float(share), 4),
        }
        for cat, share in spend_dist.items()
    ]

    # 2) 用户覆盖：每个品类覆盖多少独立用户（占全体用户比例）
    total_users = max(1, len(features))
    user_by_cat = cleaned.groupby("normalized_category", observed=True)["user_id"].nunique()
    user_dist = (
        user_by_cat.reindex(spend_by_cat.index)
        .dropna()
        .sort_values(ascending=False)
        .head(6)
    )
    category_user_distribution = [
        {
            "category": str(cat),
            "category_cn": category_display(cat),
            "user_count": int(user_by_cat[cat]),
            "user_share": round(int(user_by_cat[cat]) / total_users, 4),
        }
        for cat in user_dist.index
    ]

    # 3) Top 品类（综合加权贡献排序，取消费贡献与用户覆盖的交集优先）
    top_categories = [
        {
            "category": str(item["category"]),
            "category_cn": item["category_cn"],
            "spend_share": item["spend_share"],
            "user_share": next(
                (u["user_share"] for u in category_user_distribution if u["category"] == item["category"]),
                0.0,
            ),
        }
        for item in category_spend_distribution[:4]
    ]

    # 4) 近期活动模式
    ref_date = cleaned["event_time"].max().normalize() + pd.Timedelta(days=1)
    days_since = (ref_date - cleaned["event_time"].max()).days
    # 最近 30 天内的消费事件占比（反映"近期活跃"）
    recent_cut = cleaned["event_time"].max() - pd.Timedelta(days=30)
    recent_events = cleaned[cleaned["event_time"] >= recent_cut]
    recent_share = round(float(len(recent_events) / len(cleaned)), 4)
    # 时段分布
    hour_profile = cleaned["hour"].value_counts(normalize=True)
    peak_hours = (
        hour_profile.sort_values(ascending=False)
        .head(3)
        .index.tolist()
    )
    # 高客单品类识别：单笔均值最高的品类
    avg_amount_by_cat = cleaned.groupby("normalized_category", observed=True)["amount"].mean()
    top_ticket_cat = (
        avg_amount_by_cat.sort_values(ascending=False).head(1).index.tolist()[0]
        if not avg_amount_by_cat.empty
        else None
    )

    recent_activity_pattern = {
        "last_event_days_ago": int(days_since),
        "recent_30d_event_share": recent_share,
        "peak_hours": [int(h) for h in peak_hours],
        "high_ticket_category": category_display(top_ticket_cat) if top_ticket_cat else None,
        "high_ticket_category_value": round(float(avg_amount_by_cat.max()), 2) if not avg_amount_by_cat.empty else None,
    }

    # 5) 一句话洞察（基于真实统计拼接，非模板套话）
    top_spend_cn = category_spend_distribution[0]["category_cn"] if category_spend_distribution else "—"
    top_user_cn = category_user_distribution[0]["category_cn"] if category_user_distribution else "—"
    summary = (
        f"该人群整体消费贡献最高的是 {top_spend_cn}（金额占比 "
        f"{category_spend_distribution[0]['spend_share']:.0%}）；"
        f"覆盖用户最广的是 {top_user_cn}（占 {category_user_distribution[0]['user_share']:.0%}）。"
    )

    return {
        "available": True,
        "summary": summary,
        "top_categories": top_categories,
        "category_spend_distribution": category_spend_distribution,
        "category_user_distribution": category_user_distribution,
        "recent_activity_pattern": recent_activity_pattern,
        "segment_count": len(segments),
    }


# ---------------------------------------------------------------------------
# 人群机会排序（Segment Opportunity Ranking）
# ---------------------------------------------------------------------------
# 设计原则：
#   - 完全基于已有统计数据，不预测价值、不人工打分。
#   - 综合：人群规模、消费能力、消费贡献、活跃程度、品类集中度。
#   - 输出 opportunity_score（0~1，商业运营优先级）、opportunity_level（高/中/低）、
#     opportunity_reason（引用 statistics 的具体原因，禁止模板化、禁止"置信度/准确率"）。
# ---------------------------------------------------------------------------
def build_segment_opportunity(
    stats: dict[str, Any],
    *,
    share: float,
    max_share: float,
    segment_total_spend: float,
    overall_total_spend: float,
) -> dict[str, Any]:
    """依据真实统计对单个人群做商业机会评分。

    返回：
      {
        "opportunity_score": float,        # 0~1
        "opportunity_level": "high"|"medium"|"low",
        "opportunity_reason": [str, ...],  # 每条引用 statistics
      }
    """
    overall_monetary = float(stats.get("overall_average_spend", 0.0)) or 1.0
    monetary = float(stats.get("average_spend", 0.0))
    recency = float(stats.get("average_recency", 0.0))
    overall_recency = float(stats.get("overall_average_recency", 0.0)) or 1.0
    main_ratio = float(stats.get("main_category_ratio", 0.0) or 0.0)
    main_category = stats.get("main_category") or stats.get("dominant_category")
    main_category_cn = category_display(str(main_category)) if main_category else "单一品类"
    avg_order_value = float(stats.get("avg_order_value", 0.0) or 0.0)

    # 1) 规模：相对最大人群归一（占比越高覆盖价值越高）
    size_score = min(share / max_share, 1.0) if max_share > 0 else 0.0

    # 2) 消费能力：平均消费相对整体的倍数，clamp 到 [0, 1.5] 后归一
    spend_ratio = monetary / overall_monetary
    spend_score = max(0.0, min(spend_ratio / 1.5, 1.0))
    if avg_order_value > 0:
        aov_ratio = avg_order_value / (overall_monetary / 1.0)
        spend_score = max(spend_score, max(0.0, min(aov_ratio / 1.5, 1.0)))

    # 3) 消费贡献：该人群贡献整体消费金额的真实占比
    contribution_score = (
        float(segment_total_spend) / float(overall_total_spend)
        if overall_total_spend > 0
        else 0.0
    )
    contribution_score = max(0.0, min(contribution_score, 1.0))

    # 4) 活跃程度：平均最近消费间隔越小越活跃
    activity_score = 1.0 - max(0.0, min(recency / overall_recency, 1.0)) if overall_recency > 0 else 0.0

    # 5) 品类集中度：主品类占比越高，营销定位越容易
    concentration_score = max(0.0, min(main_ratio, 1.0))

    # 加权综合（权重之和=1）
    weights = {
        "size": 0.20,
        "spend": 0.20,
        "contribution": 0.25,
        "activity": 0.20,
        "concentration": 0.15,
    }
    score = (
        size_score * weights["size"]
        + spend_score * weights["spend"]
        + contribution_score * weights["contribution"]
        + activity_score * weights["activity"]
        + concentration_score * weights["concentration"]
    )
    score = round(max(0.0, min(score, 1.0)), 2)
    level = "high" if score >= 0.7 else ("medium" if score >= 0.45 else "low")

    reasons: list[str] = []
    reasons.append(
        f"用户规模较大（{int(stats.get('user_count', 0))} 人，占全体 {share:.0%}）"
        if share >= max_share * 0.6
        else f"人群规模 {int(stats.get('user_count', 0))} 人（占全体 {share:.0%}）"
    )
    if contribution_score >= 0.1:
        reasons.append(f"贡献整体消费金额约 {contribution_score:.0%}")
    if recency <= overall_recency * 0.8:
        reasons.append(f"最近消费平均 {recency:.0f} 天前，活跃度高于整体")
    if main_ratio >= 0.35:
        if main_ratio >= 0.99:
            reasons.append(f"消费全部集中在{main_category_cn}单一品类，定位极清晰")
        else:
            reasons.append(f"消费主要集中在{main_category_cn}品类（占 {main_ratio:.0%}），定位清晰")
    if not reasons:
        reasons.append("各项指标处于整体中位水平")
    return {
        "opportunity_score": score,
        "opportunity_level": level,
        "opportunity_reason": reasons,
    }


# ---------------------------------------------------------------------------
# 新的主分群：基于用户真实消费品类（category_weighted）进行客户分群。
# 不再使用 KMeans 作为 segment 来源。
# ---------------------------------------------------------------------------
def segment_by_category(
    cleaned: pd.DataFrame,
    features: pd.DataFrame,
    *,
    min_user_ratio: float = 0.02,
) -> tuple[pd.DataFrame, list[dict[str, Any]], dict[str, Any]]:
    """按用户主品类生成客户群体。

    流程：
      1. 计算每个用户的主品类（金额×行为权重加权最高）。
      2. 统计每个主品类下的用户数与加权贡献。
      3. 低于 total_users * min_user_ratio 的细分类别合并到其父级（一级）品类。
      4. 按主品类归群：群内复用 build_cluster_statistics 计算真实统计。
      5. 人群名称遵循 35% 阈值规则（build_category_segment_name）。

    关键约束：
      - segment.share = 真实用户数 / 总用户数，禁止人为平均分配。
      - segment_id 由主品类决定，与 KMeans 无关。
    """
    result = features.copy()
    overall_monetary = float(features["monetary"].mean())
    overall_frequency = float(features["frequency"].mean())
    overall_recency = float(features["recency"].mean())
    total_users = max(1, len(result))

    main = user_main_category(features)
    result["main_category_raw"] = main  # 原始最细粒度主品类（用于 ratio 计算）
    result["main_category_user"] = main  # 归群用（可能被合并到父级）

    # 合并过小品类到父级（一级）
    main_counts = main.value_counts()
    merge_to_parent: dict[str, str] = {}
    for cat, cnt in main_counts.items():
        if cat is None:
            continue
        if cnt < total_users * min_user_ratio:
            parent = _parent_category(cat)
            if parent:
                merge_to_parent[cat] = parent
    if merge_to_parent:
        result["main_category_user"] = result["main_category_user"].map(
            lambda c: merge_to_parent.get(c, c)
        )

    # 若合并后仍有极小父级，且其自身也不达标（例如父级仅 1 人），归入"多品类"集合
    final_counts = result["main_category_user"].value_counts()
    tiny = {c for c, cnt in final_counts.items() if cnt < total_users * min_user_ratio and c is not None}
    if tiny:
        result["main_category_user"] = result["main_category_user"].where(
            ~result["main_category_user"].isin(tiny), other=None
        )

    segments: list[dict[str, Any]] = []
    grouped = result.groupby("main_category_user", dropna=False)
    # 用户级加权矩阵（用于计算"主品类占该用户全部消费"的真实占比）
    weighted_cols_all = [c for c in result.columns if c.endswith("_weighted")]
    wm_all = result[weighted_cols_all].copy()
    wm_all.columns = [c[: -len("_weighted")] for c in wm_all.columns]
    user_total_weight = wm_all.sum(axis=1).replace(0, np.nan)
    # 逐用户：原始最细粒度主品类加权 / 该用户全部品类加权之和
    raw_series = result["main_category_raw"]
    user_main_share = pd.Series(0.0, index=result.index)
    valid_mask = raw_series.notna() & raw_series.isin(wm_all.columns)
    mc_vals = raw_series[valid_mask]
    main_weight = pd.Series(0.0, index=result.index)
    for cat in mc_vals.unique():
        idx = mc_vals[mc_vals == cat].index
        main_weight.loc[idx] = wm_all.loc[idx, cat]
    user_main_share[valid_mask] = (main_weight[valid_mask] / user_total_weight[valid_mask]).fillna(0.0)

    for grp_key, group in grouped:
        # 组内主品类加权贡献（用于命名阈值）
        weighted_cols = [c for c in group.columns if c.endswith("_weighted")]
        wm = group[weighted_cols].copy()
        wm.columns = [c[: -len("_weighted")] for c in wm.columns]
        cat_totals = wm.sum(axis=0)
        grp_top_cat = cat_totals.idxmax() if cat_totals.sum() > 0 else None
        # 主品类占比（阈值用）：组内用户"主品类/全部消费"的平均占比（用户级真实占比）。
        grp_user_shares = user_main_share.loc[group.index]
        main_ratio = round(float(grp_user_shares.mean()), 4) if grp_key is not None else 0.0
        # 主品类占比（阈值用）：组内用户"主品类/全部消费"的平均占比（用户级真实占比）。
        grp_user_shares = user_main_share.loc[group.index]
        main_ratio = round(float(grp_user_shares.mean()), 4)

        stats = build_cluster_statistics(
            group,
            overall_monetary=overall_monetary,
            overall_frequency=overall_frequency,
            overall_recency=overall_recency,
        )
        # 用组层面的真实主品类覆盖（优先使用最细粒度原始 key 的加权）
        stats["main_category"] = grp_top_cat
        stats["main_category_ratio"] = main_ratio
        stats["dominant_category"] = grp_top_cat
        stats["dominant_category_ratio"] = main_ratio
        # 人群级收入画像（输入属性，真实统计）
        seg_income = group["income"].dropna() if "income" in group.columns else pd.Series(dtype=float)
        stats["average_income"] = float(seg_income.mean()) if not seg_income.empty else None

        # 判断是否全部来自 electronics 一级（用于聚合命名）
        if grp_key is None or grp_top_cat is None:
            name = "综合兴趣用户"
            electronics_only = False
            grp_top_cat = grp_top_cat  # 可能为 None
        else:
            top_level = grp_top_cat.split(".")[0] if "." in str(grp_top_cat) else str(grp_top_cat)
            electronics_only = top_level == "electronics"
            name = build_category_segment_name(
                grp_top_cat, main_ratio, electronics_only=electronics_only, stats=stats
            )
        share = len(group) / total_users

        evidence = [
            {
                "metric": "用户规模",
                "value": f"{len(group)} 人（{share:.1%}）",
                "benchmark": f"全体 {total_users} 人",
                "interpretation": "该人群由主品类相同的真实用户构成，占比为用户数/总用户数。",
            },
            {
                "metric": "主品类加权贡献",
                "value": (
                    f"{category_display(str(grp_top_cat))} "
                    f"（占该人群加权消费 {main_ratio:.0%}）"
                ),
                "benchmark": "金额×行为权重",
                "interpretation": "按金额×行为权重（购买>加购>浏览）计算的人群主品类贡献占比。",
            },
            {
                "metric": "平均消费",
                "value": f"¥{stats['average_spend']:.0f}",
                "benchmark": f"整体 ¥{overall_monetary:.0f}（{f"{stats['spend_ratio']:.1f}" if stats['spend_ratio'] >= 0.1 else f"{stats['spend_ratio']:.2f}"} 倍）",
                "interpretation": "人群内人均消费与整体对比，仅描述样本内相对贡献。",
            },
            {
                "metric": "平均购买频次",
                "value": f"{stats['average_frequency']:.1f} 次",
                "benchmark": f"整体 {overall_frequency:.1f} 次",
                "interpretation": "人群的复购或互动节奏。",
            },
        ]

        seg_income_profile = build_income_profile(group)
        segments.append(
            {
                "segment_id": f"cat::{grp_key}",
                "cluster_id": None,
                "name": name,
                "user_count": int(len(group)),
                "share": round(share, 4),
                "income_profile": seg_income_profile,
                "statistics": stats,
                "key_features": build_cluster_profile(stats)[1],
                "evidence": evidence,
                "recommended_strategy": build_recommended_strategy(stats),
                "rule_basis": [
                    "按用户真实消费品类（金额×行为权重）确定主品类，再据此归群。",
                    "人群占比 = 真实用户数 / 总用户数，不做人为平均分配。",
                    "单一主品类占比 < 35% 时不使用单一商品类命名。",
                    "少于总用户 1%~3% 的细分类别合并到父级品类。",
                    "收入仅作为画像字段，不参与人群筛选。",
                ],
            }
        )

    # ---- 人群机会排序：基于已有统计计算商业运营优先级（不预测、不人工打分）----
    overall_total_spend = float(result["total_purchase_amount"].sum())
    max_share = max((item["share"] for item in segments), default=0.0)
    for seg in segments:
        seg_stats = seg.get("statistics", {}) or {}
        seg_total_spend = float(seg_stats.get("total_purchase_amount", 0.0) or 0.0)
        opportunity = build_segment_opportunity(
            seg_stats,
            share=seg["share"],
            max_share=max_share,
            segment_total_spend=seg_total_spend,
            overall_total_spend=overall_total_spend,
        )
        seg["opportunity_score"] = opportunity["opportunity_score"]
        seg["opportunity_level"] = opportunity["opportunity_level"]
        seg["opportunity_reason"] = opportunity["opportunity_reason"]
        seg_stats.update(opportunity)

    segments.sort(key=lambda item: item["user_count"], reverse=True)
    segment_quality: dict[str, Any] = {
        "method": "category_preference",
        "segment_count": len(segments),
        "sample_size": int(total_users),
        "min_user_ratio": min_user_ratio,
        "note": "基于用户真实消费品类的客户分群；KMeans 仅作为群内价值分层辅助分析。",
    }
    return result, segments, segment_quality


def segment_users(features: pd.DataFrame, cluster_count: int | None = None) -> tuple[pd.DataFrame, list[dict[str, Any]], dict[str, Any]]:
    """保留接口兼容：当前主分群已改为 segment_by_category。

    此函数仅用于「群内价值分层」辅助分析（KMeans），不再决定 segment_id。
    """
    result = features.copy()
    numeric_columns = list(CLUSTERING_FEATURES)
    cluster_quality: dict[str, Any] = {
        "silhouette_score": None,
        "cluster_count": int(cluster_count or settings.default_cluster_count),
        "sample_size": int(len(result)),
        "search_range": [2, 8],
        "clustering_features": list(CLUSTERING_FEATURES),
        "role": "auxiliary_value_tier",
    }

    if len(result) >= 2 and result[numeric_columns].nunique().sum() > 0:
        matrix = StandardScaler().fit_transform(result[numeric_columns])
        if cluster_count:
            n_clusters = max(2, min(cluster_count, len(result)))
        else:
            n_clusters, _ = select_best_cluster_count(matrix)
        model = KMeans(n_clusters=n_clusters, random_state=42, n_init=10)
        result["cluster_id"] = model.fit_predict(matrix)
        try:
            score = silhouette_score(matrix, result["cluster_id"])
        except ValueError:
            score = None
        cluster_quality["cluster_count"] = int(n_clusters)
        cluster_quality["silhouette_score"] = score
    else:
        result["cluster_id"] = 0
        cluster_quality["note"] = "样本不足或特征无方差，辅助分层降级为单组。"
    return result, [], cluster_quality


@dataclass
class DataToolResult:
    quality: dict[str, Any]
    cleaned: pd.DataFrame | None
    features: pd.DataFrame | None
    segments: list[dict[str, Any]]
    cleaning_stats: dict[str, int]
    cluster_quality: dict[str, Any]
    segment_method: str = "category_preference"
    category_debug: dict[str, Any] = field(default_factory=dict)
    category_warning: str | None = None
    income_profile: dict[str, Any] = field(default_factory=dict)
    overall_consumption_insight: dict[str, Any] = field(default_factory=dict)


def build_category_debug(raw: pd.DataFrame, cleaned: pd.DataFrame, mapping: dict[str, str]) -> tuple[dict[str, Any], str | None]:
    """生成品类调试信息，验证原始类别 -> 标准类别是否发生异常归并。

    返回 (category_debug, category_warning)：
      - raw_top_categories: 原始 category 取值 Top10（含占比）
      - normalized_top_categories: 标准化后 normalized_category Top10（含占比）
      - mapping_changes: 发生变化的 (raw -> normalized, count) Top10
      - category_warning: 若单一标准化品类占比 > 90%，给出提示（不修改数据）。
    """
    debug: dict[str, Any] = {
        "raw_top_categories": [],
        "normalized_top_categories": [],
        "mapping_changes": [],
    }
    warning: str | None = None
    if "category" not in mapping:
        return debug, warning

    raw_col = mapping["category"]
    total = len(cleaned)
    if total == 0:
        return debug, warning

    # 原始 Top10
    raw_counts = raw[raw_col].astype(str).fillna("").replace({"nan": ""})
    raw_top = raw_counts[raw_counts != ""].value_counts().head(10)
    debug["raw_top_categories"] = [
        {"category": str(k), "count": int(v), "share": round(float(v) / total, 4)}
        for k, v in raw_top.items()
    ]

    # 标准化 Top10
    norm_top = cleaned["normalized_category"].value_counts().head(10)
    debug["normalized_top_categories"] = [
        {"category": str(k), "count": int(v), "share": round(float(v) / total, 4)}
        for k, v in norm_top.items()
    ]

    # 发生变化的映射（raw -> normalized 不一致）
    paired = pd.DataFrame(
        {
            "raw": raw[raw_col].astype(str).fillna(""),
            "normalized": cleaned["normalized_category"].astype(str),
        }
    )
    changed = paired[paired["raw"].str.strip() != paired["normalized"]]
    change_counts = changed.groupby(["raw", "normalized"]).size().sort_values(ascending=False).head(10)
    debug["mapping_changes"] = [
        {"raw": str(r), "normalized": str(n), "count": int(c)}
        for (r, n), c in change_counts.items()
    ]

    # 异常检测：单一标准化品类占比 > 90%
    top_norm_share = float(norm_top.iloc[0] / total) if len(norm_top) else 0.0
    if top_norm_share > 0.9:
        warning = "单一品类占比过高，请检查类别标准化逻辑"
    return debug, warning


def analyze_file(path: str | Path) -> DataToolResult:
    raw = read_dataset(path)
    quality = build_quality_report(raw)
    if not quality["can_analyze"]:
        return DataToolResult(quality, None, None, [], {}, {}, {}, None)
    cleaned, stats = clean_dataset(raw, quality["field_mapping"])
    if cleaned.empty:
        quality["can_analyze"] = False
        quality["issues"].append(
            {
                "code": "no_valid_rows",
                "severity": "error",
                "message": "清洗后没有可用记录。",
                "field": None,
            }
        )
        return DataToolResult(quality, cleaned, None, [], stats, {}, {}, None)
    features = build_user_features(cleaned)
    _, segments, cluster_quality = segment_by_category(cleaned, features)
    # 保留 KMeans 仅作辅助价值分层（不影响 segment_id）
    features, _, _ = segment_users(features)
    category_debug, category_warning = build_category_debug(raw, cleaned, quality["field_mapping"])
    income_profile = build_income_profile(features)
    consumption_insight = build_consumption_trend(cleaned, features, segments)
    return DataToolResult(
        quality,
        cleaned,
        features,
        segments,
        stats,
        cluster_quality,
        segment_method=cluster_quality.get("method", "category_preference"),
        category_debug=category_debug,
        category_warning=category_warning,
        income_profile=income_profile,
        overall_consumption_insight=consumption_insight,
    )


def build_upload_preview(raw: pd.DataFrame, total_rows: int) -> tuple[dict[str, Any], dict[str, Any]]:
    quality = build_quality_report(raw)
    quality["row_count"] = total_rows
    quality["preview_row_count"] = len(raw)
    if total_rows > len(raw):
        quality["issues"].append(
            {
                "code": "sampled_preview",
                "severity": "info",
                "message": f"上传阶段基于前 {len(raw):,} 行快速预览；点击开始分析后会读取完整数据。",
                "field": None,
            }
        )
    preview: dict[str, Any] = {
        "sample_row_count": len(raw),
        "segments": [],
        "insights": [],
        "top_products": [],
        "date_range": "",
        "avg_order_value": 0.0,
        "total_amount": 0.0,
        "income_profile": {},
        "overall_consumption_insight": {},
    }
    if not quality["can_analyze"]:
        return quality, preview

    cleaned, _ = clean_dataset(raw, quality["field_mapping"])
    if cleaned.empty:
        return quality, preview

    features = build_user_features(cleaned)
    _, segments, _ = segment_by_category(cleaned, features)
    preview["segments"] = segments[:4]
    preview["income_profile"] = build_income_profile(features)
    preview["overall_consumption_insight"] = build_consumption_trend(cleaned, features, segments)
    preview["date_range"] = (
        f"{cleaned['event_time'].min().date()} 至 {cleaned['event_time'].max().date()}"
    )
    preview["avg_order_value"] = round(float(cleaned["amount"].mean()), 2)
    preview["total_amount"] = round(float(cleaned["amount"].sum()), 2)
    preview["top_products"] = [
        {"name": str(name), "count": int(count)}
        for name, count in cleaned["product"].value_counts().head(5).items()
    ]
    preview["insights"] = [
        {
            "title": segment["name"],
            "summary": "；".join(segment["key_features"][:2]),
            "evidence": [item["value"] for item in segment["evidence"][:3]],
        }
        for segment in segments[:4]
    ]
    return quality, preview
